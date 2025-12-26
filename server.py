from flask import Flask, request, jsonify, send_file, make_response
from flask_cors import CORS
from datetime import datetime, timedelta
import json
import os
import hashlib
import uuid
import threading
import time
import pandas as pd
from io import BytesIO
import base64
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

app = Flask(__name__)
# CORS для работы с любого домена
CORS(app, supports_credentials=True, origins=["*"], 
     allow_headers=["Content-Type", "Authorization"],
     methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"])

# Конфигурация
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'cargo-secret-key-2024-abu-1610')
app.config['DATABASE_URL'] = os.environ.get('DATABASE_URL', 'database.json')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# Для Render.com или другого хостинга
if os.environ.get('RENDER'):
    app.config['DATA_FOLDER'] = '/opt/render/project/src/data'
    app.config['UPLOAD_FOLDER'] = '/opt/render/project/src/uploads'
else:
    app.config['DATA_FOLDER'] = 'data'
    app.config['UPLOAD_FOLDER'] = 'uploads'

# Создаем папки
os.makedirs(app.config['DATA_FOLDER'], exist_ok=True)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Путь к файлу базы данных
DB_FILE = os.path.join(app.config['DATA_FOLDER'], app.config['DATABASE_URL'])

# Блокировка для синхронизации между запросами
db_lock = threading.Lock()

# Кэш базы данных в памяти
db_cache = None
db_last_modified = 0

# ==================== БАЗА ДАННЫХ ====================

def init_db():
    """Инициализирует базу данных"""
    global db_cache, db_last_modified
    
    if not os.path.exists(DB_FILE):
        default_db = {
            'users': [
                {
                    'id': str(uuid.uuid4()),
                    'email': '12abc202@gmail.com',
                    'password': hashlib.sha256('Abu_1610'.encode()).hexdigest(),
                    'isAdmin': True,
                    'createdAt': datetime.now().isoformat(),
                    'lastSync': datetime.now().isoformat(),
                    'devices': []
                }
            ],
            'orders': [],
            'settings': {
                'kgz_rate': 88.5,
                'company_name': 'Cargo Management System',
                'company_phone': '+996 XXX XXX XXX'
            },
            'sessions': [],
            'sync_log': [],
            'last_modified': datetime.now().isoformat()
        }
        save_db(default_db)
        db_cache = default_db
        db_last_modified = time.time()
    return load_db()

def load_db():
    """Загружает базу данных"""
    global db_cache, db_last_modified
    
    try:
        # Проверяем кэш
        if db_cache and os.path.exists(DB_FILE):
            file_modified = os.path.getmtime(DB_FILE)
            if file_modified <= db_last_modified:
                return db_cache.copy()
        
        # Загружаем из файла
        with db_lock:
            with open(DB_FILE, 'r', encoding='utf-8') as f:
                db_cache = json.load(f)
                db_last_modified = time.time()
                return db_cache.copy()
    except:
        return init_db()

def save_db(data):
    """Сохраняет базу данных"""
    global db_cache, db_last_modified
    
    try:
        with db_lock:
            data['last_modified'] = datetime.now().isoformat()
            
            with open(DB_FILE, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            db_cache = data.copy()
            db_last_modified = time.time()
            
        return True
    except Exception as e:
        print(f"Error saving DB: {e}")
        return False

def generate_token():
    return str(uuid.uuid4())

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def verify_token(token):
    """Проверяет токен"""
    if not token:
        return None
    
    db = load_db()
    sessions = db.get('sessions', [])
    
    # Очищаем старые сессии
    active_sessions = []
    user = None
    
    for session in sessions:
        try:
            created = datetime.fromisoformat(session['created'])
            if datetime.now() - created < timedelta(days=30):
                active_sessions.append(session)
                if session['token'] == token:
                    user_id = session.get('user_id')
                    for u in db.get('users', []):
                        if u['id'] == user_id:
                            user = u
                            # Обновляем последнюю активность
                            u['lastActivity'] = datetime.now().isoformat()
                            break
        except:
            continue
    
    if len(active_sessions) != len(sessions):
        db['sessions'] = active_sessions
        save_db(db)
    
    return user

# ==================== АВТОРИЗАЦИЯ ====================

@app.route('/api/auth/register', methods=['POST'])
def register():
    try:
        data = request.json
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        
        if not email or not password:
            return jsonify({'error': 'Заполните все поля'}), 400
        
        db = load_db()
        
        for user in db.get('users', []):
            if user['email'] == email:
                return jsonify({'error': 'Пользователь уже существует'}), 400
        
        user_id = str(uuid.uuid4())
        new_user = {
            'id': user_id,
            'email': email,
            'password': hash_password(password),
            'isAdmin': email == '12abc202@gmail.com',
            'createdAt': datetime.now().isoformat(),
            'lastSync': datetime.now().isoformat(),
            'lastActivity': datetime.now().isoformat(),
            'devices': []
        }
        
        db['users'].append(new_user)
        
        session_token = generate_token()
        db['sessions'].append({
            'token': session_token,
            'user_id': user_id,
            'created': datetime.now().isoformat()
        })
        
        save_db(db)
        
        return jsonify({
            'token': session_token,
            'user': {
                'id': user_id,
                'email': email,
                'isAdmin': new_user['isAdmin'],
                'uid': email
            }
        }), 201
        
    except Exception as e:
        print(f"Register error: {str(e)}")
        return jsonify({'error': 'Ошибка сервера'}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    try:
        data = request.json
        email = data.get('email', '').strip().lower()
        password = data.get('password', '')
        device_id = data.get('deviceId', str(uuid.uuid4()))
        device_name = data.get('deviceName', request.user_agent.string[:50] if request.user_agent else 'Unknown')
        
        if not email or not password:
            return jsonify({'error': 'Заполните все поля'}), 400
        
        db = load_db()
        
        user = None
        for u in db.get('users', []):
            if u['email'] == email and u['password'] == hash_password(password):
                user = u
                break
        
        if not user:
            return jsonify({'error': 'Неверные учетные данные'}), 401
        
        # Регистрируем устройство
        if 'devices' not in user:
            user['devices'] = []
        
        device_found = False
        for device in user['devices']:
            if device.get('deviceId') == device_id:
                device['lastSeen'] = datetime.now().isoformat()
                device['deviceName'] = device_name
                device_found = True
                break
        
        if not device_found:
            user['devices'].append({
                'deviceId': device_id,
                'deviceName': device_name,
                'firstSeen': datetime.now().isoformat(),
                'lastSeen': datetime.now().isoformat()
            })
        
        user['lastSync'] = datetime.now().isoformat()
        user['lastActivity'] = datetime.now().isoformat()
        
        # Обновляем пользователя
        for i, u in enumerate(db['users']):
            if u['id'] == user['id']:
                db['users'][i] = user
                break
        
        # Создаем сессию
        session_token = generate_token()
        db['sessions'].append({
            'token': session_token,
            'user_id': user['id'],
            'device_id': device_id,
            'created': datetime.now().isoformat()
        })
        
        save_db(db)
        
        return jsonify({
            'token': session_token,
            'user': {
                'id': user['id'],
                'email': user['email'],
                'isAdmin': user.get('isAdmin', False),
                'uid': user['email']
            },
            'syncInfo': {
                'lastSync': user.get('lastSync'),
                'devicesCount': len(user.get('devices', [])),
                'serverTime': datetime.now().isoformat()
            }
        })
        
    except Exception as e:
        print(f"Login error: {str(e)}")
        return jsonify({'error': 'Ошибка сервера'}), 500

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    
    if token:
        db = load_db()
        db['sessions'] = [s for s in db.get('sessions', []) if s['token'] != token]
        save_db(db)
    
    return jsonify({'success': True})

@app.route('/api/auth/check', methods=['GET'])
def check_auth():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if user:
        return jsonify({
            'authenticated': True,
            'user': {
                'id': user['id'],
                'email': user['email'],
                'isAdmin': user.get('isAdmin', False),
                'uid': user['email']
            }
        })
    
    return jsonify({'authenticated': False}), 401

# ==================== ЗАКАЗЫ ====================

@app.route('/api/orders', methods=['GET'])
def get_orders():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    db = load_db()
    orders = db.get('orders', [])
    
    # Админ видит все, остальные - только свои
    if not user.get('isAdmin', False):
        orders = [o for o in orders if o.get('userId') == user['id']]
    
    return jsonify({
        'orders': orders,
        'total': len(orders),
        'lastSync': db.get('last_modified')
    })

@app.route('/api/orders', methods=['POST'])
def create_order():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    
    order_id = str(uuid.uuid4())
    new_order = {
        'id': order_id,
        'userId': user['id'],
        'userEmail': user['email'],
        'fileName': data.get('fileName', ''),
        'uploadDate': data.get('uploadDate', datetime.now().strftime('%d.%m.%Y')),
        'data': data.get('data', []),
        'createdAt': datetime.now().isoformat(),
        'updatedAt': datetime.now().isoformat(),
        'syncVersion': 1
    }
    
    db = load_db()
    
    # Проверка на дубликат
    for existing_order in db.get('orders', []):
        if (existing_order.get('fileName') == new_order['fileName'] and 
            existing_order.get('userId') == user['id']):
            # Обновляем существующий
            existing_order['data'] = new_order['data']
            existing_order['updatedAt'] = datetime.now().isoformat()
            existing_order['syncVersion'] = existing_order.get('syncVersion', 0) + 1
            save_db(db)
            return jsonify({'order': existing_order, 'updated': True})
    
    db['orders'].append(new_order)
    save_db(db)
    
    return jsonify({'order': new_order, 'created': True}), 201

@app.route('/api/orders/<order_id>', methods=['PUT'])
def update_order(order_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    db = load_db()
    
    for i, order in enumerate(db.get('orders', [])):
        if order['id'] == order_id:
            if not user.get('isAdmin') and order.get('userId') != user['id']:
                return jsonify({'error': 'Forbidden'}), 403
            
            order.update(data)
            order['updatedAt'] = datetime.now().isoformat()
            order['syncVersion'] = order.get('syncVersion', 0) + 1
            db['orders'][i] = order
            save_db(db)
            
            return jsonify({'order': order})
    
    return jsonify({'error': 'Order not found'}), 404

@app.route('/api/orders/<order_id>', methods=['DELETE'])
def delete_order(order_id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    db = load_db()
    orders = db.get('orders', [])
    
    for i, order in enumerate(orders):
        if order['id'] == order_id:
            if not user.get('isAdmin') and order.get('userId') != user['id']:
                return jsonify({'error': 'Forbidden'}), 403
            
            orders.pop(i)
            db['orders'] = orders
            save_db(db)
            
            return jsonify({'success': True, 'message': 'Order deleted'})
    
    return jsonify({'error': 'Order not found'}), 404

# ==================== ЭКСПОРТ В EXCEL ====================

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.json
        filter_type = data.get('filter', 'all')
        
        db = load_db()
        orders = db.get('orders', [])
        
        if not user.get('isAdmin'):
            orders = [o for o in orders if o.get('userId') == user['id']]
        
        # Создаем новый Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'База данных'
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = ['№', 'Рейс', 'Дата', 'Имя', 'Вес (кг)', 'Сумма ($)', 
                  'Доплата', 'Юл Кира', 'Итог ($)', 'Сом', 'Сотрудник', 'Статус']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        row_num = 2
        total_sum = 0
        total_paid = 0
        total_unpaid = 0
        
        for order in orders:
            order_data = order.get('data', [])
            if len(order_data) <= 1:
                continue
                
            for row_data in order_data[1:]:  # Пропускаем заголовки
                status = row_data[11] if len(row_data) > 11 else 'не оплачено'
                
                # Фильтрация
                if filter_type == 'paid' and status != 'оплачено':
                    continue
                elif filter_type == 'unpaid' and status == 'оплачено':
                    continue
                
                # Расчеты
                sum_dollar = float(row_data[5]) if len(row_data) > 5 and row_data[5] else 0
                doplata = float(row_data[6]) if len(row_data) > 6 and row_data[6] else 0
                yul_kira = float(row_data[7]) if len(row_data) > 7 and row_data[7] else 0
                total = sum_dollar + doplata + yul_kira
                som = (sum_dollar + yul_kira) * 88.5
                
                total_sum += total
                if status == 'оплачено':
                    total_paid += total
                else:
                    total_unpaid += total
                
                # Заполняем строку
                ws.cell(row=row_num, column=1, value=row_num-1).border = border
                ws.cell(row=row_num, column=2, value=order.get('fileName', '')).border = border
                ws.cell(row=row_num, column=3, value=row_data[1] if len(row_data) > 1 else '').border = border
                ws.cell(row=row_num, column=4, value=row_data[2] if len(row_data) > 2 else '').border = border
                ws.cell(row=row_num, column=5, value=row_data[4] if len(row_data) > 4 else '').border = border
                ws.cell(row=row_num, column=6, value=sum_dollar).border = border
                ws.cell(row=row_num, column=7, value=doplata).border = border
                ws.cell(row=row_num, column=8, value=yul_kira).border = border
                ws.cell(row=row_num, column=9, value=total).border = border
                ws.cell(row=row_num, column=10, value=som).border = border
                ws.cell(row=row_num, column=11, value=row_data[10] if len(row_data) > 10 else '').border = border
                
                status_cell = ws.cell(row=row_num, column=12, value=status)
                status_cell.border = border
                if status == 'оплачено':
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row_num += 1
        
        # Итоги
        row_num += 1
        ws.cell(row=row_num, column=1, value='ИТОГО:').font = Font(bold=True)
        ws.cell(row=row_num, column=9, value=f'${total_sum:.2f}').font = Font(bold=True)
        
        row_num += 1
        ws.cell(row=row_num, column=1, value='Оплачено:').font = Font(bold=True, color="008000")
        ws.cell(row=row_num, column=9, value=f'${total_paid:.2f}').font = Font(bold=True, color="008000")
        
        row_num += 1
        ws.cell(row=row_num, column=1, value='Не оплачено:').font = Font(bold=True, color="FF0000")
        ws.cell(row=row_num, column=9, value=f'${total_unpaid:.2f}').font = Font(bold=True, color="FF0000")
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем в память
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'cargo_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        return jsonify({'error': 'Ошибка экспорта'}), 500

# ==================== ЭКСПОРТ В PDF ====================

@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.json
        filter_type = data.get('filter', 'all')
        
        db = load_db()
        orders = db.get('orders', [])
        settings = db.get('settings', {})
        
        if not user.get('isAdmin'):
            orders = [o for o in orders if o.get('userId') == user['id']]
        
        # Создаем PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Стили
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER
        )
        
        # Заголовок
        company_name = settings.get('company_name', 'Cargo Management System')
        elements.append(Paragraph(company_name, title_style))
        elements.append(Spacer(1, 20))
        
        # Информация
        info_style = styles['Normal']
        elements.append(Paragraph(f'Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")}', info_style))
        elements.append(Paragraph(f'Пользователь: {user["email"]}', info_style))
        elements.append(Spacer(1, 20))
        
        # Таблица данных
        table_data = [['№', 'Рейс', 'Дата', 'Имя', 'Сумма ($)', 'Статус']]
        
        total_sum = 0
        total_paid = 0
        total_unpaid = 0
        row_num = 1
        
        for order in orders:
            order_data = order.get('data', [])
            if len(order_data) <= 1:
                continue
                
            for row_data in order_data[1:]:
                status = row_data[11] if len(row_data) > 11 else 'не оплачено'
                
                if filter_type == 'paid' and status != 'оплачено':
                    continue
                elif filter_type == 'unpaid' and status == 'оплачено':
                    continue
                
                sum_dollar = float(row_data[5]) if len(row_data) > 5 and row_data[5] else 0
                doplata = float(row_data[6]) if len(row_data) > 6 and row_data[6] else 0
                yul_kira = float(row_data[7]) if len(row_data) > 7 and row_data[7] else 0
                total = sum_dollar + doplata + yul_kira
                
                total_sum += total
                if status == 'оплачено':
                    total_paid += total
                else:
                    total_unpaid += total
                
                table_data.append([
                    str(row_num),
                    order.get('fileName', ''),
                    row_data[1] if len(row_data) > 1 else '',
                    row_data[2] if len(row_data) > 2 else '',
                    f'${total:.2f}',
                    status
                ])
                row_num += 1
        
        # Создаем таблицу
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Итоги
        summary_data = [
            ['Итого:', f'${total_sum:.2f}'],
            ['Оплачено:', f'${total_paid:.2f}'],
            ['Не оплачено:', f'${total_unpaid:.2f}']
        ]
        
        summary_table = Table(summary_data, colWidths=[100, 100])
        summary_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(summary_table)
        
        # Генерируем PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'cargo_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        )
        
    except Exception as e:
        print(f"PDF export error: {str(e)}")
        return jsonify({'error': 'Ошибка экспорта PDF'}), 500

# ==================== СИНХРОНИЗАЦИЯ ====================

@app.route('/api/sync', methods=['POST'])
def sync_data():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        data = request.json
        client_version = data.get('version', 0)
        client_orders = data.get('orders', [])
        
        db = load_db()
        server_orders = db.get('orders', [])
        
        # Фильтруем заказы пользователя
        if not user.get('isAdmin'):
            server_orders = [o for o in server_orders if o.get('userId') == user['id']]
        
        # Простая синхронизация: берем более новые версии
        merged_orders = {}
        
        # Добавляем серверные заказы
        for order in server_orders:
            order_id = order['id']
            merged_orders[order_id] = order
        
        # Мержим с клиентскими
        for order in client_orders:
            order_id = order.get('id')
            if not order_id:
                # Новый заказ с клиента
                order_id = str(uuid.uuid4())
                order['id'] = order_id
                order['userId'] = user['id']
                order['userEmail'] = user['email']
                order['syncVersion'] = 1
                merged_orders[order_id] = order
            else:
                # Существующий заказ
                if order_id in merged_orders:
                    server_order = merged_orders[order_id]
                    # Сравниваем версии
                    client_ver = order.get('syncVersion', 0)
                    server_ver = server_order.get('syncVersion', 0)
                    
                    if client_ver > server_ver:
                        merged_orders[order_id] = order
                else:
                    # Заказ есть на клиенте, но нет на сервере
                    order['userId'] = user['id']
                    order['userEmail'] = user['email']
                    merged_orders[order_id] = order
        
        # Сохраняем обновленные заказы
        all_orders = db.get('orders', [])
        user_order_ids = set(merged_orders.keys())
        
        # Удаляем старые заказы пользователя
        all_orders = [o for o in all_orders if o.get('userId') != user['id'] or o['id'] not in user_order_ids]
        
        # Добавляем обновленные
        all_orders.extend(merged_orders.values())
        
        db['orders'] = all_orders
        save_db(db)
        
        return jsonify({
            'success': True,
            'orders': list(merged_orders.values()),
            'serverTime': datetime.now().isoformat(),
            'syncVersion': db.get('last_modified')
        })
        
    except Exception as e:
        print(f"Sync error: {str(e)}")
        return jsonify({'error': 'Ошибка синхронизации'}), 500

# ==================== НАСТРОЙКИ ====================

@app.route('/api/settings/<key>', methods=['GET'])
def get_setting(key):
    db = load_db()
    settings = db.get('settings', {})
    
    if key in settings:
        return jsonify({'value': settings[key]})
    
    defaults = {
        'kgz_rate': 88.5,
        'company_name': 'Cargo Management System',
        'company_phone': '+996 XXX XXX XXX'
    }
    
    if key in defaults:
        return jsonify({'value': defaults[key]})
    
    return jsonify({'error': 'Setting not found'}), 404

@app.route('/api/settings/<key>', methods=['POST'])
def set_setting(key):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    data = request.json
    value = data.get('value')
    
    db = load_db()
    if 'settings' not in db:
        db['settings'] = {}
    
    db['settings'][key] = value
    save_db(db)
    
    return jsonify({'success': True, 'key': key, 'value': value})

# ==================== СТАТИСТИКА ====================

@app.route('/api/stats', methods=['GET'])
def get_stats():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user = verify_token(token)
    
    if not user:
        return jsonify({'error': 'Unauthorized'}), 401
    
    db = load_db()
    orders = db.get('orders', [])
    
    if not user.get('isAdmin'):
        orders = [o for o in orders if o.get('userId') == user['id']]
    
    total_orders = len(orders)
    total_sum = 0
    total_paid = 0
    total_unpaid = 0
    employee_stats = {}
    
    for order in orders:
        order_data = order.get('data', [])
        for row in order_data[1:]:
            if len(row) > 11:
                sum_val = float(row[5]) if row[5] else 0
                doplata = float(row[6]) if row[6] else 0
                yul_kira = float(row[7]) if row[7] else 0
                total = sum_val + doplata + yul_kira
                
                total_sum += total
                
                if row[11] == 'оплачено':
                    total_paid += total
                else:
                    total_unpaid += total
                
                employee = row[10] if len(row) > 10 else 'Unknown'
                if employee:
                    if employee not in employee_stats:
                        employee_stats[employee] = {'paid': 0, 'unpaid': 0}
                    
                    if row[11] == 'оплачено':
                        employee_stats[employee]['paid'] += total
                    else:
                        employee_stats[employee]['unpaid'] += total
    
    return jsonify({
        'totalOrders': total_orders,
        'totalSum': total_sum,
        'totalPaid': total_paid,
        'totalUnpaid': total_unpaid,
        'employeeStats': employee_stats,
        'lastSync': db.get('last_modified')
    })

# ==================== HEALTH CHECK ====================

@app.route('/api/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '2.0.0',
        'sync_enabled': True,
        'pdf_export': True,
        'excel_export': True
    })

@app.route('/')
def index():
    return jsonify({
        'name': 'Cargo Management API',
        'version': '2.0.0',
        'status': 'running',
        'features': [
            'Multi-device sync',
            'PDF export',
            'Excel export',
            'Real-time updates'
        ],
        'endpoints': {
            'auth': ['/api/auth/login', '/api/auth/register', '/api/auth/logout'],
            'orders': ['/api/orders', '/api/orders/<id>'],
            'export': ['/api/export/pdf', '/api/export/excel'],
            'sync': ['/api/sync'],
            'stats': ['/api/stats']
        }
    })

# ==================== ЗАПУСК ====================

if __name__ == '__main__':
    init_db()
    
    # Для продакшена используем переменные окружения
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('DEBUG', 'True').lower() == 'true'
    
    app.run(
        host='0.0.0.0',
        port=port,
        debug=debug
    )
